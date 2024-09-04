import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

def export_to_excel(results, file_path, platform):
    try:
        logger.info(f"Starting export to Excel: {file_path}")
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if isinstance(results, dict) and 'product' in results:
            # Product Info Fetcher results
            sheet = workbook.create_sheet(title=f"{platform} Product Info")
            products = [p for p in results['product'] if p is not None]
            if products:
                export_product_info(sheet, products, platform, timestamp)
            else:
                sheet.cell(row=1, column=1, value="No valid product information found")
        else:
            # Rank Fetcher results
            for keyword, products in results.items():
                sheet = workbook.create_sheet(title=keyword[:31])  # Excel sheet names limited to 31 characters
                if products:
                    export_rank_fetcher_results(sheet, products, platform, timestamp)
                else:
                    sheet.cell(row=1, column=1, value=f"No products found for '{keyword}'")

        logger.info(f"Saving workbook to: {file_path}")
        workbook.save(file_path)
        logger.info(f"Results exported successfully to {file_path}")
    except Exception as e:
        logger.error(f"Error exporting results to Excel: {str(e)}")
        raise
    finally:
        logger.info("Excel export operation completed")

def export_product_info(sheet, products, platform, timestamp):
    if platform == "Amazon":
        headers = ["S.No", "ASIN", "Link", "Title", "Price", "Rating", "Reviews", "BestSeller", "In Stock", "Timestamp"]
    else:  # Flipkart
        headers = ["S.No", "Product ID", "Link", "Title", "Price", "Rating", "Reviews", "Timestamp"]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, product in enumerate(products, start=2):
        sheet.cell(row=row, column=1, value=row-1)  # S.No
        if platform == "Amazon":
            sheet.cell(row=row, column=2, value=product.get("ASIN", "N/A"))
            sheet.cell(row=row, column=3, value=product.get("link", "N/A"))
            sheet.cell(row=row, column=4, value=product.get("title", "N/A"))
            sheet.cell(row=row, column=5, value=product.get("price", "N/A"))
            sheet.cell(row=row, column=6, value=product.get("rating", "N/A"))
            sheet.cell(row=row, column=7, value=product.get("reviews", "N/A"))
            sheet.cell(row=row, column=8, value=product.get("BestSeller", "N/A"))
            sheet.cell(row=row, column=9, value=product.get("In Stock", "N/A"))
            sheet.cell(row=row, column=10, value=timestamp)
        else:  # Flipkart
            sheet.cell(row=row, column=2, value=product.get("product_id", "N/A"))
            sheet.cell(row=row, column=3, value=product.get("link", "N/A"))
            sheet.cell(row=row, column=4, value=product.get("title", "N/A"))
            sheet.cell(row=row, column=5, value=product.get("price", "N/A"))
            sheet.cell(row=row, column=6, value=product.get("rating", "N/A"))
            sheet.cell(row=row, column=7, value=product.get("reviews", "N/A"))
            sheet.cell(row=row, column=8, value=timestamp)

def export_rank_fetcher_results(sheet, products, platform, timestamp):
    if platform == "Amazon":
        headers = ["Rank", "ASIN", "Link", "Title", "Price", "Rating", "Reviews", "Type", "Timestamp"]
    else:  # Flipkart
        headers = ["Rank", "Product ID", "Link", "Title", "Price", "Rating", "Reviews", "Timestamp"]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, product in enumerate(products, start=2):
        col = 1
        sheet.cell(row=row, column=col, value=product.get("rank", "N/A")); col += 1
        sheet.cell(row=row, column=col, value=product.get("asin") or product.get("product_id", "N/A")); col += 1
        sheet.cell(row=row, column=col, value=product.get("link", "N/A")); col += 1
        sheet.cell(row=row, column=col, value=product.get("title", "N/A")); col += 1
        sheet.cell(row=row, column=col, value=product.get("price", "N/A")); col += 1
        sheet.cell(row=row, column=col, value=product.get("rating", "N/A")); col += 1
        sheet.cell(row=row, column=col, value=product.get("reviews", "N/A")); col += 1
        
        if platform == "Amazon":
            sheet.cell(row=row, column=col, value=product.get("type", "N/A")); col += 1
        
        sheet.cell(row=row, column=col, value=timestamp)